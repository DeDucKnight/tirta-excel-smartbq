from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtWidgets import QMessageBox, QFileDialog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import pandas as pd
import os
import re
from collections import defaultdict

class Ui_MainWindow(object):    
    def setupUi(self, MainWindow):
        # MainWindow setup
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 600)
        self.centralwidget = QtWidgets.QWidget(parent=MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.formLayoutWidget = QtWidgets.QWidget(parent=self.centralwidget)
        self.formLayoutWidget.setGeometry(QtCore.QRect(20, 10, 761, 321))
        self.formLayoutWidget.setObjectName("formLayoutWidget")
        self.formLayout = QtWidgets.QFormLayout(self.formLayoutWidget)
        self.formLayout.setLabelAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.formLayout.setContentsMargins(0, 0, 0, 0)
        self.formLayout.setObjectName("formLayout")
        self.uploadButton = QtWidgets.QPushButton(parent=self.formLayoutWidget)
        self.uploadButton.setObjectName("uploadButton")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.ItemRole.LabelRole, self.uploadButton)
        self.fileNameLabel = QtWidgets.QLabel(parent=self.formLayoutWidget)
        self.fileNameLabel.setText("")
        self.fileNameLabel.setObjectName("fileNameLabel")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.ItemRole.FieldRole, self.fileNameLabel)

        # Analisa Sheet
        self.analisaSheetLabel = QtWidgets.QLabel(parent=self.formLayoutWidget)
        self.analisaSheetLabel.setObjectName("analisaSheetLabel")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.ItemRole.LabelRole, self.analisaSheetLabel)
        self.analisaSheetComboBox = QtWidgets.QComboBox(parent=self.formLayoutWidget)
        self.analisaSheetComboBox.setObjectName("analisaSheetComboBox")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.ItemRole.FieldRole, self.analisaSheetComboBox)

        # Analisa Beton Sheet
        self.analisaBetonSheetLabel = QtWidgets.QLabel(parent=self.formLayoutWidget)
        self.analisaBetonSheetLabel.setLayoutDirection(QtCore.Qt.LayoutDirection.LeftToRight)
        self.analisaBetonSheetLabel.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.analisaBetonSheetLabel.setObjectName("analisaBetonSheetLabel")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.ItemRole.LabelRole, self.analisaBetonSheetLabel)
        self.analisaBetonSheetComboBox = QtWidgets.QComboBox(parent=self.formLayoutWidget)
        self.analisaBetonSheetComboBox.setObjectName("analisaBetonSheetComboBox")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.ItemRole.FieldRole, self.analisaBetonSheetComboBox)

        # Target Sheet
        self.targetSheetLabel = QtWidgets.QLabel(parent=self.formLayoutWidget)
        self.targetSheetLabel.setObjectName("targetSheetLabel")
        self.formLayout.setWidget(3, QtWidgets.QFormLayout.ItemRole.LabelRole, self.targetSheetLabel)
        self.targetListWidget = QtWidgets.QListWidget(parent=self.formLayoutWidget)
        self.targetListWidget.setObjectName("targetListWidget")
        self.targetListWidget.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.MultiSelection)
        self.formLayout.setWidget(3, QtWidgets.QFormLayout.ItemRole.FieldRole, self.targetListWidget)

        # Generate Button
        self.generateButton = QtWidgets.QPushButton(parent=self.centralwidget)
        self.generateButton.setGeometry(QtCore.QRect(680, 350, 101, 41))
        self.generateButton.setObjectName("generateButton")


        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(parent=MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 22))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(parent=MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.uploadButton.clicked.connect(self.on_uploadButton_clicked)
        self.generateButton.clicked.connect(self.on_generateButton_clicked)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.uploadButton.setText(_translate("MainWindow", "Upload"))
        self.analisaSheetLabel.setText(_translate("MainWindow", "Sheet Analisa"))
        self.analisaBetonSheetLabel.setText(_translate("MainWindow", "Sheet Beton"))
        self.targetSheetLabel.setText(_translate("MainWindow", "Sheet Target"))
        self.generateButton.setText(_translate("MainWindow", "Generate"))

    def ref_to_idx(self, cell_ref):
        m = re.match(r"([A-Za-z]+)(\d+)", cell_ref)
        if not m:
            return None, None
        col_letters, row_str = m.groups()
        row_idx = int(row_str) - 1
        # Convert column letters to index (A=0, B=1, ...)
        col_idx = sum((ord(c.upper()) - ord('A') + 1) * (26 ** i) for i, c in enumerate(col_letters[::-1])) - 1
        return row_idx, col_idx
    
    def calculate_formula(self, formula):
        try:
            if isinstance(formula, str) and formula.startswith('='):
                expression = formula.lstrip('=').strip()
                if re.fullmatch(r"[0-9.*+/() \t-]+", expression):  # basic safety check
                    return eval(expression)
                else:
                    return 0.0  # fallback if not safe
            else:
                return float(formula) if formula is not None else 0.0
        except Exception as e:
            return 0.0

    def preprocess_ws(self, ws):
        def resolve(cell_row, col_index, visited=None):
            if visited is None:
                visited = set()
            key = (cell_row, col_index)
            if key in visited:
                return ws.cell(row=cell_row, column=col_index).value  # prevent infinite loop
            visited.add(key)

            val = ws.cell(row=cell_row, column=col_index).value
            if not isinstance(val, str) or not val.startswith('='):
                return val

            ref = val.lstrip('=').strip()

            # Local cell reference like E10 or D7
            m = re.match(r"([A-Z]+)(\d+)$", ref, re.IGNORECASE)
            if m:
                target_row, target_col = self.ref_to_idx(ref)
                if target_row is not None:
                    return resolve(target_row + 1, target_col + 1, visited)

            # Arithmetic expression like =1*0.6
            if re.fullmatch(r"[0-9.*+/() \\t-]+", ref):
                try:
                    return eval(ref)
                except:
                    return val

            # Cross-sheet or unsupported formula → return as-is
            return "=" + ref

        for row in range(2, ws.max_row + 1):
            for col in [4, 5]:
                formula = ws.cell(row=row, column=col).value
                if isinstance(formula, str) and formula.startswith('='):
                    final = resolve(row, col)
                    ws.cell(row=row, column=col).value = final

    def resolve_formula(self, sheet_dict, sheet_name, row, col, visited=None):
        if visited is None:
            visited = set()
        key = (sheet_name, row, col)
        if key in visited:
            return None  # circular ref guard
        visited.add(key)

        ws = sheet_dict.get(sheet_name)
        if not ws:
            return None

        cell = ws.cell(row=row + 1, column=col + 1)
        val = cell.value

        if not isinstance(val, str) or not val.startswith("="):
            return val

        ref = val.lstrip("=").strip()

        # Match arithmetic expressions including cell references
        if re.fullmatch(r"[A-Za-z0-9\.\+\-\*/\(\) \t]+", ref):
            try:
                # Find all cell references like D3, E8, etc.
                tokens = re.findall(r"\b([A-Za-z]+[0-9]+)\b", ref)
                for token in set(tokens):  # set avoids duplicates
                    r, c = self.ref_to_idx(token)
                    if r is not None:
                        val = self.resolve_formula(sheet_dict, sheet_name, r, c, visited.copy())
                        ref = ref.replace(token, str(val))
                return eval(ref)
            except:
                return 0.0
            
        if ref.upper().startswith("SUM(") and ":" in ref:
            try:
                range_str = ref[4:-1]
                start_cell, end_cell = range_str.split(":")
                start_row, start_col = self.ref_to_idx(start_cell)
                end_row, end_col = self.ref_to_idx(end_cell)
                total = 0.0
                for r in range(start_row, end_row + 1):
                    val = self.resolve_formula(sheet_dict, sheet_name, r, start_col, visited.copy())
                    try:
                        total += float(val)
                    except:
                        continue
                return total
            except:
                return 0.0
            
        # Reference to another cell
        if "!" in ref:
            if ref.startswith("'"):
                tgt_sheet, cell_ref = ref.split("'!")
                tgt_sheet = tgt_sheet.strip("'")
            else:
                tgt_sheet, cell_ref = ref.split("!")
        else:
            tgt_sheet = sheet_name  # local reference
            cell_ref = ref

        tgt_row, tgt_col = self.ref_to_idx(cell_ref)
        if tgt_row is None:
            return val

        return self.resolve_formula(sheet_dict, tgt_sheet, tgt_row, tgt_col, visited)
    
    def resolve_formula_target_sheet(self, sheet_dict, sheet_name, row, col, visited=None):
        if visited is None:
            visited = set()
        key = (sheet_name, row, col)
        if key in visited:
            return None  # prevent circular loop
        visited.add(key)

        ws = sheet_dict.get(sheet_name)
        if not ws:
            return None

        cell = ws.cell(row=row + 1, column=col + 1)
        val = cell.value

        if not isinstance(val, str) or not val.startswith("="):
            return val

        ref = val.lstrip("=").strip()

        # Arithmetic-only expression
        if re.fullmatch(r"[0-9\.\+\-\*/\(\) \t]+", ref):
            try:
                return eval(ref)
            except:
                return 0.0

        # Check if it's a reference
        if "!" in ref:
            if ref.startswith("'"):
                tgt_sheet, cell_ref = ref.split("'!")
                tgt_sheet = tgt_sheet.strip("'")
            else:
                tgt_sheet, cell_ref = ref.split("!")
        else:
            tgt_sheet = sheet_name
            cell_ref = ref

        # Skip if it's Analisa or Beton
        analisa_sheet = self.analisaSheetComboBox.currentText()
        beton_sheet = self.analisaBetonSheetComboBox.currentText()
        if tgt_sheet.lower() == analisa_sheet.lower() or tgt_sheet.lower() == beton_sheet.lower():
            return val  # return the formula as-is

        tgt_row, tgt_col = self.ref_to_idx(cell_ref)
        if tgt_row is None:
            return val

        return self.resolve_formula_target_sheet(sheet_dict, tgt_sheet, tgt_row, tgt_col, visited)

    def resolve_local_formula(self, sheet_dict, sheet_name, row, col, visited=None):
        if visited is None:
            visited = set()
        key = (sheet_name, row, col)
        if key in visited:
            return None  # prevent circular reference
        visited.add(key)

        ws = sheet_dict.get(sheet_name)
        if not ws:
            return None

        cell = ws.cell(row=row + 1, column=col + 1)
        val = cell.value

        if not isinstance(val, str) or not val.startswith("="):
            return val

        ref = val.lstrip("=").strip()

        # Stop on cross-sheet reference
        if "!" in ref:
            return val

        # Stop on expression or math (we're only handling single refs)
        if re.search(r"[+\-*/()]", ref):
            return val

        # Only handle simple self-referencing like "=E10"
        target_row, target_col = self.ref_to_idx(ref)
        if target_row is None:
            return val

        return self.resolve_local_formula(sheet_dict, sheet_name, target_row, target_col, visited)

    def resolve_formula_v2(self, sheet_dict, sheet_name, row, col, visited=None, stop_on_cross_sheet=False, exclude_sheets=None):
        if visited is None:
            visited = set()
        if exclude_sheets is None:
            exclude_sheets = []

        key = (sheet_name, row, col)
        if key in visited:
            return None  # prevent circular reference
        visited.add(key)

        ws = sheet_dict.get(sheet_name)
        if not ws:
            return None

        val = ws.cell(row=row + 1, column=col + 1).value

        if not isinstance(val, str) or not val.startswith("="):
            return val

        ref = val.lstrip("=").strip()

        # Cross-sheet check
        if "!" in ref:
            if ref.startswith("'"):
                tgt_sheet, cell_ref = ref.split("'!")
                tgt_sheet = tgt_sheet.strip("'")
            else:
                tgt_sheet, cell_ref = ref.split("!")
            if stop_on_cross_sheet or tgt_sheet.lower() in [s.lower() for s in exclude_sheets]:
                return val
        else:
            tgt_sheet = sheet_name
            cell_ref = ref
        
        # Handle arithmetic expression with cell references
        if re.fullmatch(r"[A-Za-z0-9\.\+\-\*/\(\) \t]+", ref):
            try:
                tokens = re.findall(r"\b([A-Za-z]+[0-9]+)\b", ref)
                resolved_ref = ref
                for token in set(tokens):
                    r, c = self.ref_to_idx(token)
                    resolved = self.resolve_formula_v2(sheet_dict, sheet_name, r, c, visited.copy(), stop_on_cross_sheet, exclude_sheets)
                    if isinstance(resolved, (int, float)):
                        resolved_ref = re.sub(rf"\b{token}\b", str(resolved), resolved_ref)
                    elif isinstance(resolved, str):
                        resolved_ref = re.sub(rf"\b{token}\b", f'"{resolved}"', resolved_ref)
                    else:
                        return resolved
                return eval(resolved_ref)
            except:
                return 0.0

        # Handle SUM(...)
        if ref.upper().startswith("SUM(") and ":" in ref:
            try:
                range_str = ref[4:-1]
                start_cell, end_cell = range_str.split(":")
                start_row, start_col = self.ref_to_idx(start_cell)
                end_row, end_col = self.ref_to_idx(end_cell)
                total = 0.0
                for r in range(start_row, end_row + 1):
                    val = self.resolve_formula_v2(sheet_dict, sheet_name, r, start_col, visited.copy(), stop_on_cross_sheet, exclude_sheets)
                    if isinstance(val, (int, float)):
                        total += val
                return total
            except:
                return 0.0

        tgt_row, tgt_col = self.ref_to_idx(cell_ref)
        if tgt_row is None:
            return val

        return self.resolve_formula_v2(sheet_dict, tgt_sheet, tgt_row, tgt_col, visited.copy(), stop_on_cross_sheet, exclude_sheets)

    def extract_volume_rows_from_ws(self, ws, formula, multiplier=1.0, sheet_dict=None, analisa_ws=None):
        result = []

        if not formula.startswith("="):
            return result

        ref = formula.lstrip("=").strip()
        if ref.startswith("'"):
            sheet, cell = ref.split("'!")
            sheet = sheet.strip("'")
        else:
            sheet, cell = ref.split("!")

        r_idx, c_idx = self.ref_to_idx(cell)
        if r_idx is None:
            return result

        # Find the last 'T' above the reference row
        t_row = None
        for i in range(r_idx - 1, -1, -1):
            val = ws.cell(row=i + 1, column=c_idx + 1).value
            if str(val).strip().upper() == 'T':
                t_row = i + 1
                break

        if t_row is None:
            return result

        for i in range(t_row, r_idx):
            try:
                name = ws.cell(row=i + 1, column=2).value  # Column B
                unit = ws.cell(row=i + 1, column=3).value  # Column C
                coef = ws.cell(row=i + 1, column=4).value  # Column D
                price_ref = self.resolve_local_formula(sheet_dict, sheet, i, 4) if sheet_dict else 0.0

                if isinstance(name, str) and name.strip().startswith("-"):
                    name = name.strip().lstrip("-").strip()

                clean_coef = self.calculate_formula(coef)
                price = self.resolve_formula_v2(sheet_dict, sheet, i, 4) if sheet_dict else 0.0

                # Price Subtotal
                price_subtotal = clean_coef * price

                # Append the beton material itself
                result.append({
                    "name": name,
                    "unit": unit,
                    "volume": clean_coef * multiplier,
                    "price": price,
                    "price_subtotal": price_subtotal,
                    "row": i
                })

                # Check if price_ref points to Analisa
                if isinstance(price_ref, str) and 'analisa' in price_ref.lower() and analisa_ws:
                    sub_materials = self.extract_volume_rows_from_ws(
                        analisa_ws,
                        price_ref,
                        multiplier=clean_coef * multiplier,
                        sheet_dict=sheet_dict,
                        analisa_ws=analisa_ws
                    )
                    result.extend(sub_materials)

            except Exception:
                continue

        return result
    
    def classify_formula(self, ref: str, analisa_name: str, beton_name: str) -> str:
        if not isinstance(ref, str) or not ref.startswith("="):
            return "simple"

        # Normalize for easier comparison
        ref = ref.strip().lstrip('=').lower()

        has_analisa = analisa_name.lower() in ref
        has_beton = beton_name.lower() in ref
        has_operator = any(op in ref for op in ['+', '-', '*', '/', '(', ')'])

        if (has_analisa or has_beton) and has_operator:
            return "mixed"
        elif has_analisa or has_beton:
            return "analisa_only"
        else:
            return "simple"

    def on_uploadButton_clicked(self):
        options = QFileDialog.Option.DontUseNativeDialog
        file_filter = "Excel Files (*.xlsx *.xls)"
        file_name, _ = QFileDialog.getOpenFileName(None, "Select Excel File", "", file_filter, options=options)

        if file_name:
            self.file_path = file_name
            self.fileNameLabel.setText(file_name)
            try:
                xls = pd.ExcelFile(file_name)
                sheet_names = xls.sheet_names
                self.populate_comboboxes(sheet_names)
            except Exception as e:
                QMessageBox.critical(None, "Error", f"Failed to read file:\n{str(e)}")

    def populate_comboboxes(self, sheet_names):
        self.analisaSheetComboBox.clear()
        self.analisaBetonSheetComboBox.clear()

        self.targetListWidget.clear()
        self.targetListWidget.addItems(sheet_names)

        analisa_found = False
        for sheet in sheet_names:
            if sheet.lower() == "analisa":
                self.analisaSheetComboBox.addItem(sheet)
                analisa_found = True
                break
        if not analisa_found:
            self.analisaSheetComboBox.addItems(sheet_names)

        beton_found = False
        for sheet in sheet_names:
            if sheet.upper().startswith("AN BETON "):
                self.analisaBetonSheetComboBox.addItem(sheet)
                beton_found = True
                break
        if not beton_found:
            self.analisaBetonSheetComboBox.addItems(sheet_names)

    def on_generateButton_clicked(self):
        try:
            if not hasattr(self, "file_path"):
                QMessageBox.warning(None, "Warning", "Please upload an Excel file first.")
                return

            wb = load_workbook(self.file_path, data_only=False)
            analisa_name = self.analisaSheetComboBox.currentText()
            beton_name = self.analisaBetonSheetComboBox.currentText()
            analisa_ws = wb[analisa_name]
            beton_ws = wb[beton_name]

            # Preprocess Analisa and Beton sheets
            self.preprocess_ws(analisa_ws)
            self.preprocess_ws(beton_ws)

            target_sheet_names = [item.text() for item in self.targetListWidget.selectedItems()]
            sheet_dict = {name: wb[name] for name in target_sheet_names}
            sheet_dict[analisa_name] = analisa_ws
            sheet_dict[beton_name] = beton_ws

            all_results = {}            
            all_total_materials = {}

            for sheet_name in target_sheet_names:
                results = []
                total_materials = defaultdict(lambda: {"volume": 0.0, "price_subtotal": 0.0})
                ws = sheet_dict[sheet_name]
                debug = "5. Buang tanah bekas galian pondasi"
                for row in range(2, ws.max_row + 1):
                    name = ws.cell(row=row, column=2).value
                    if debug in str(name):
                        print(f"Debugging row {row} in {sheet_name}: {name}")
                    volume = self.resolve_formula_v2(sheet_dict, sheet_name, row - 1, 2)
                    price = self.resolve_formula_v2(sheet_dict, sheet_name, row - 1, 4)
                    formula = self.resolve_formula_v2(sheet_dict, sheet_name, row - 1, 4, exclude_sheets=[self.analisaSheetComboBox.currentText(), self.analisaBetonSheetComboBox.currentText()])

                    try:
                        volume = float(volume)
                    except:
                        continue

                    results.append({"item": name, "material": "", "volume": volume, "formula": formula, "price": price, "unit": ws.cell(row=row, column=4).value})

                    if not isinstance(formula, str) or not formula.startswith("="):
                        continue

                    ref = formula.lstrip("=").strip()
                    if "!" in ref:
                        if ref.startswith("'"):
                            sheet, cell = ref.split("'!")
                            sheet = sheet.strip("'")
                        else:
                            sheet, cell = ref.split("!")
                    else:
                        sheet = sheet_name
                        cell = ref

                    if sheet.lower() == analisa_name.lower(): # Analisa Sheet
                        sub_items = self.extract_volume_rows_from_ws(
                            sheet_dict[sheet],
                            formula,
                            multiplier=volume,
                            sheet_dict=sheet_dict,
                            analisa_ws=analisa_ws
                        )
                        for item in sub_items:
                            results.append({
                                    "item": "",
                                    "material": item["name"],
                                    "volume": item["volume"],
                                    "formula": "",
                                    "unit": item.get("unit", ""),
                                    "price": item.get("price", 0.0)
                                })
                            price_subtotal = item["volume"] * item.get("price", 0.0)
                            total_materials[item["name"]]["volume"] += item["volume"]
                            total_materials[item["name"]]["price_subtotal"] += price_subtotal
                    elif sheet.lower() == beton_name.lower(): # Beton Sheet
                        sub_items = self.extract_volume_rows_from_ws(
                            sheet_dict[sheet],
                            formula,
                            multiplier=volume,
                            sheet_dict=sheet_dict,
                            analisa_ws=analisa_ws
                        )
                        
                        for item in sub_items:
                            # If sub is a beton material (top-level), it has both name and price
                            if item.get("price") is not None:
                                results.append({
                                    "item": "",
                                    "material": item["name"],
                                    "volume": item["volume"],
                                    "formula": "",
                                    "unit": item.get("unit", ""),
                                    "price": item.get("price", 0.0),
                                    "price_subtotal": item.get("price_subtotal", 0.0)
                                })
                                price_subtotal = item["volume"] * item.get("price", 0.0)
                                total_materials[item["name"]]["volume"] += item["volume"]
                                total_materials[item["name"]]["price_subtotal"] += price_subtotal
                            # If it's a deeper Analisa item already extracted
                            else:
                                results.append({
                                    "item": "",
                                    "material": item["name"],
                                    "volume": item["volume"],
                                    "formula": "",
                                    "unit": item.get("unit", ""),
                                    "price": item.get("price", 0.0),
                                    "price_subtotal": item.get("price_subtotal", 0.0)
                                })
                                price_subtotal = item["volume"] * item.get("price", 0.0)
                                total_materials[item["name"]]["volume"] += item["volume"]
                                total_materials[item["name"]]["price_subtotal"] += price_subtotal
                
                all_results[sheet_name] = results
                all_total_materials[sheet_name] = total_materials
            
            wb_out = Workbook()
            if 'Sheet' in wb_out.sheetnames:
                del wb_out['Sheet']

            # Raw Material Sheets
            for sheet_name, rows in all_results.items():
                ws_raw = wb_out.create_sheet(title=f"Raw - {sheet_name[:28]}")
                ws_raw.append(["Items", "Materials", "Volume", "Unit", "Price"])

                ws_raw.column_dimensions['A'].width = 90
                ws_raw.column_dimensions['B'].width = 90
                for row in rows:
                    row_idx = ws_raw.max_row + 1
                    ws_raw.cell(row=row_idx, column=1, value=row.get("item", ""))
                    ws_raw.cell(row=row_idx, column=2, value=row.get("material", ""))
                    ws_raw.cell(row=row_idx, column=3, value=round(row.get("volume", 0.0), 2))
                    ws_raw.cell(row=row_idx, column=4, value=row.get("unit", ""))
                    ws_raw.cell(row=row_idx, column=5, value=row.get("price", 0.0))

                    formula = row.get("formula", "")
                    formula_type = self.classify_formula(formula, analisa_name, beton_name)
                    if formula_type in ("simple", "mixed"):
                        ws_raw.cell(row=row_idx, column=1).font = Font(color="FF0000")
                        
            # Total Material Sheet
            for sheet_name, total_materials in all_total_materials.items():
                ws_total = wb_out.create_sheet(title=f"Total - {sheet_name[:28]}")
                ws_total.append(["Materials", "Total Volume", "Total Price"])
                ws_total.column_dimensions['A'].width = 90

                for name, total in total_materials.items():
                    volume = total["volume"]
                    subtotal = total["price_subtotal"]
                    ws_total.append([name, round(volume, 4), round(subtotal, 4)])

            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            default = os.path.join(desktop, "FormattedOutput.xlsx")
            path, _ = QFileDialog.getSaveFileName(None, "Save Excel File", default, "Excel Files (*.xlsx)")
            if path:
                wb_out.save(path)
                QMessageBox.information(None, "Success", f"File saved to:\n{path}")

        except Exception as e:
            QMessageBox.critical(None, "Error", f"An error occurred:\n{str(e)}")

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec())